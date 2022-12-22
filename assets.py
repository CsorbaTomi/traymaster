import openpyxl, os, datetime

# first step create the data save file rewrite to test enough use once
def create_data_file(file_name):
    wb = openpyxl.Workbook()
    wb.create_sheet('Gepek')
    wb.create_sheet('Downtime')
    gepek = wb["Gepek"]
    gepek.append(["Dátum", "Berendezés", "I szalag", "II szalag", "Bevitte"])
    gepek.append([dt(), "Test", "0", "", "Testman"])
    downtime = wb["Downtime"]
    downtime.append(["Gep", "Név", "Időpont", "OK", "Időtartam"])
    downtime.append(["Test", "Testman", dt(), "first setting", 0])
    wb.save(f"{file_name}.xlsx")
    workbook = openpyxl.load_workbook(f"{file_name}.xlsx")
    std = workbook["Sheet"]
    workbook.remove(std)
    workbook.save(f"{file_name}.xlsx")
# get user
def login():
    return os.getlogin()
# get date
def date():
    return datetime.datetime.strptime(datetime.datetime.now().strftime("%Y-%m-%d"),"%Y-%m-%d")
# get time0
def time():
    return datetime.datetime.now().strftime("%H:%M:%S")
# get datetime
def dt():
    return datetime.datetime.strptime(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),"%Y-%m-%d %H:%M:%S")
# save working data in the file    
def working(file_name, machine, user, state, date_time=0, work_time=0):
    wb = openpyxl.load_workbook(file_name)
    wb.active
    work = wb["Downtime"]
    work.append([machine, user, date_time, state, work_time])
    wb.save(file_name)
# save pause data in the file
def pause(file_name, machine, user, state, date_time=0, work_time=0):
    wb = openpyxl.load_workbook(file_name)
    wb.active
    pause = wb["Downtime"]
    pause.append([machine, user, date_time, state, work_time])
    wb.save(file_name)
    
# save cleaning data in the file
def cleaning(file_name, machine, user, state, date_time=0, work_time=0):
    wb = openpyxl.load_workbook(file_name)
    wb.active
    clean = wb["Downtime"]
    clean.append([machine, user, date_time, state, work_time])
    wb.save(file_name)
# save maintenance data in the file
def maintenance(file_name, machine, user, state, date_time=0, work_time=0):
    wb = openpyxl.load_workbook(file_name)
    wb.active
    maintenance = wb["Downtime"]
    maintenance.append([machine, user, date_time, state, work_time])
    wb.save(file_name)
# save not using data in the file
def not_using(file_name, machine, user, state, date_time=0, work_time=0):
    wb = openpyxl.load_workbook(file_name)
    wb.active
    not_use = wb["Downtime"]
    not_use.append([machine, user, date_time, state, work_time])
    wb.save(file_name)
# save shift change data in the file
def shift_change(file_name, machine, user, state, count_number, date_time=0, work_time=0):
    wb = openpyxl.load_workbook(file_name)
    wb.active
    shift_change = wb["Downtime"]
    shift_change.append([machine, user, date_time, state, work_time])
    save_counter = wb["Gepek"]
    save_counter.append([date_time, machine, count_number, "", user])
    wb.save(file_name)
# save and set the zero data in the file
def set_zero_default_value(file_name, machine, user, date_time=0):   
    wb = openpyxl.load_workbook(file_name)
    wb.active
    set_zero = wb["Gepek"]
    set_zero.append([date_time, machine, 0, "", user])
    wb.save(file_name)
# get last value
def last_value(file_name):
    wb = openpyxl.load_workbook(file_name)
    wb.active
    gepek = wb["Gepek"]
    max_row = gepek.max_row
    return gepek[f"C{max_row}"].value
# get last state
def last_state(file_name):
    wb = openpyxl.load_workbook(file_name)
    wb.active
    downtime = wb["Downtime"]
    max_row = downtime.max_row
    return downtime[f"D{max_row}"].value
# get last but one value
def last_state1(file_name):
    wb = openpyxl.load_workbook(file_name)
    wb.active
    downtime = wb["Downtime"]
    max_row = downtime.max_row
    return downtime[f"D{max_row - 1}"].value
# get last time
def last_time(file_name):
    wb = openpyxl.load_workbook(file_name)
    wb.active
    downtime = wb["Downtime"]
    max_row = downtime.max_row
    return downtime[f"C{max_row}"].value

# if __name__ == "__main__":
#